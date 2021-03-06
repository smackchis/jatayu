from openpyxl import *
from openpyxl.styles import Alignment, PatternFill
from datetime import *
import globals


ConfigWB = load_workbook(filename=globals.ConfigPath)
PathWS = ConfigWB['PATHS']
EtcWS = ConfigWB['ETC']
ProjectsWS = ConfigWB['PROJECTS']
ProjectPath = PathWS.cell(row=1, column=2).value
TestResultPath = PathWS.cell(row=2, column=2).value
WaitTime = EtcWS.cell(row=1,column=2).value
DriverPath = globals.ProjectPath + "etc\\chromedriver.exe"
wb = load_workbook(filename=globals.test_suite_path)
ws = wb[globals.test_sheet]
total_cases = int(str(ws.cell(row=ws.max_row, column=1).value).strip())
RepoSheet = wb['REPO']

def get_test_data(test_id,test_step):
    row_no = 3
    end_row  = ws.max_row
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                test_data = str(ws.cell(row=row_no, column=4).value).strip()
                return test_data

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1


def test_status(test_id,test_step,test_status):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                if test_status.lower() == 'pass':
                    ws.cell(row=row_no, column=9).fill = PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')
                if test_status.lower() == 'fail':
                    ws.cell(row=row_no, column=9).fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
                if test_status.lower() == 'pending':
                    ws.cell(row=row_no, column=9).fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
                ws.cell(row=row_no, column=9).value = test_status
                ws.cell(row=row_no, column=10).value = datetime.now()

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1
    save_test_resut(ws)


def AutomationException(test_id,test_step,exception):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                ws.cell(row=row_no, column=11).value = exception

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1

    save_test_resut(ws)


def WriteTestparameters(test_id,test_step,**kwargs):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                col=12
                for key, value in kwargs.items():
                    ws.cell(row=row_no, column=col).value = key + "=" + value
                    col = col + 1

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1
    save_test_resut(ws)

def get_element_repo(ElementName):
    row_no = 2
    end_row  = RepoSheet.max_row
    while row_no <= RepoSheet.max_row:

        if ElementName == str(RepoSheet.cell(row=row_no, column=2).value).strip():
            IdentifierType = str(RepoSheet.cell(row=row_no, column=3).value).strip()
            ElementType = str(RepoSheet.cell(row=row_no, column=4).value).strip()
            ElementIdentifier = str(RepoSheet.cell(row=row_no, column=5).value).strip()
            Action = str(RepoSheet.cell(row=row_no, column=6).value).strip()
            #print(ElementName, IdentifierType, ElementType, ElementIdentifier, Action)
            return IdentifierType, ElementType, ElementIdentifier, Action
        row_no = row_no + 1

def get_expected_result(test_id, test_step):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                expected_result = str(ws.cell(row=row_no, column=8).value).strip().split('::')[2]
                expected_result_variables = expected_result.split(' == ')[0]
                expected_result_values = expected_result.split(' == ')[1]
                expected_result_variable_list = expected_result_variables.split(';;')
                expected_values_list = expected_result_values.split(';;')
        row_no = row_no + 1
    for n in range(0,len(expected_result_variable_list)):
        expected_result_variable_list[n] = expected_result_variable_list[n].split('||')
        expected_values_list[n] = expected_values_list[n].split('||')
    return expected_result_variable_list, expected_values_list

def compare_actual_and_excepted_result(test_id, test_step, expected_result_variable_list, expected_values_list, actual_value_list):
    #print(expected_result_variable_list, expected_values_list, actual_value_list)
    test_result = True
    for n in range(len(expected_result_variable_list)):
        row_no = 3
        while row_no <= ws.max_row:
            if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
                if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                    print_text = ''
                    my_test_result = False
                    ws.cell(row=row_no, column=n+12).fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
                    for m in range(0,len(expected_result_variable_list[n])):
                        print_text = print_text + (expected_result_variable_list[n])[m] + ": EXPECTED= '" + (expected_values_list[n])[m] + "'   ACTUAL= '" + (actual_value_list[n])[m] + "'" +  '     OR' + '\n'
                        if (expected_values_list[n])[m] == (actual_value_list[n])[m]:
                            my_test_result = True
                            ws.cell(row=row_no, column=n+12).fill = PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')
                    ws.cell(row=row_no, column=n+12).alignment = Alignment(wrap_text=True)
                    ws.cell(row=row_no, column=n+12).value = print_text
                    test_result = my_test_result * test_result
                    #print(print_text)

            elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
                break
            row_no = row_no + 1

        save_test_resut(ws)
    if test_result:
        test_status(test_id,test_step,'PASS')
    else:
        test_status(test_id,test_step,'FAIL')


def save_test_resut(ws):
    wb.save(str(globals.ProjectPath) + "\\test_suite\\" + str(globals.project) + "\\" + str(globals.test_sheet) + "_test_result.xlsx")

wb.close()
