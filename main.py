from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
import importlib
import os, sys, time
from datetime import datetime
import test_sheet_handler, globals
from importlib import reload
from test_modules import drop_down, click, enter_value, get_url
from expected_result_modules import verify_value

def run_tests(test_id):
    row_no = 3
    while row_no <= test_sheet_handler.ws.max_row:
        if str(test_sheet_handler.ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            test_module = str(test_sheet_handler.ws.cell(row=row_no, column=2).value).strip()
            test_variable = str(test_sheet_handler.ws.cell(row=row_no, column=3).value).strip()
            test_step = str(test_sheet_handler.ws.cell(row=row_no, column=6).value).strip()
            test_data = str(test_sheet_handler.ws.cell(row=row_no, column=4).value)
            expected_result_module = str(test_sheet_handler.ws.cell(row=row_no, column=8).value)
            if test_sheet_handler.ws.cell(row=row_no, column=8).value != None:
                expected_result_module = (str(test_sheet_handler.ws.cell(row=row_no, column=8).value).strip().split('::')[1])[1:]

            try:
                if test_variable != None:
                    IdentifierType, ElementType, ElementIdentifier, Action = test_sheet_handler.get_element_repo(test_variable)
                    time.sleep(test_sheet_handler.WaitTime)
                    print(datetime.now(), "\t test_id:", test_id, "\t test_step:", test_step, "\t TestModule:",test_module, "\t TestElement:", test_variable, "\t test_data:", test_data, end='')
                    print("\t WebElementsREPO", "\t TestElemnt:", IdentifierType, "\t ElementType:", ElementType, "\t Identifier:", ElementIdentifier, "\t Action:", Action)
                    if test_module == 'get_url':
                        get_url.url(driver, test_id, test_step, test_data).Link()
                    if test_module == 'click':
                        click.LeftClick(driver, test_id, test_step, test_data, test_variable, IdentifierType, ElementType, ElementIdentifier, Action).ClickByElement()
                    if test_module == 'enter_value':
                        enter_value.SendValue(driver, test_id, test_step, test_data, test_variable, IdentifierType, ElementType, ElementIdentifier, Action).SendKeysByElement()
                    if test_module == 'drop_down':
                        drop_down.SelectFromDropDown(driver, test_id, test_step, test_data, test_variable, IdentifierType, ElementType, ElementIdentifier, Action).SelectElelmentType()


                    if expected_result_module == 'verify_value':
                        time.sleep(test_sheet_handler.WaitTime)
                        verify_value.CompareResults(driver, test_id, test_step, test_data, test_variable, IdentifierType, ElementType, ElementIdentifier, Action).GetTestStatus()

                    if expected_result_module == 'None':
                        test_sheet_handler.test_status(test_id,test_step,'PASS')

            except:
                print(sys.exc_info())
                exception = sys.exc_info()[0]
                print(exception)
                test_sheet_handler.AutomationException(test_id, test_step, str(exception))
                test_sheet_handler.test_status(test_id,test_step,'PENDING')
            finally:
                pass
        row_no = row_no + 1


driver = webdriver.Chrome(executable_path=test_sheet_handler.DriverPath)
driver.maximize_window()

'''
windriver = webdriver.Remote(command_executor='http://localhost:9999'
        ,desired_capabilities={"debugConnectToRunningApp": 'false'
        ,"app": r"C:\Windows\System32"})

windriver.find_element_by_name("").click()
view_menu_item = window.find_element_by_id('MenuBar').find_element_by_name('View')  
windriver.find_element_by_xpath("(//*[contains(@Name,'hhh')])[7]").click()
'''

TestSuitPath = globals.ProjectPath + "test_suite\\"

for prow in range(2,test_sheet_handler.ProjectsWS.max_row+1):
    globals.project  = str(test_sheet_handler.ProjectsWS.cell(row=prow,column=1).value).strip()
    if globals.project != 'None' and (str(test_sheet_handler.ProjectsWS.cell(row=prow,column=2).value).strip()).lower() == 'yes':
        TestSuitPath = globals.ProjectPath + "test_suite\\" + globals.project + "\\"
        for TestSuits in os.listdir(TestSuitPath):
            if TestSuits.endswith(".xlsm") and "~$" not in TestSuits:
                globals.test_suite_path = os.path.join(TestSuitPath, TestSuits)
                globals.test_sheet = TestSuits[2:-5] + '_TestCase'
                reload(test_sheet_handler)
                print(globals.test_suite_path, globals.test_sheet)
                test_id, row = 1,3
                while test_id <= int(test_sheet_handler.total_cases):
                    run_tests(test_id)
                    test_id = test_id + 1

#driver.close()
