from openpyxl import *
from datetime import *


wb = load_workbook(filename="C:\\TM\\jatayu\\test_suite\\LOGIN.xlsm")
ws = wb['LOGIN_TestCase']
RepoSheet = wb['REPO']
total_cases = int(str(ws.cell(row=ws.max_row, column=1).value).strip())


def get_test_data(test_id,test_step):
    row_no = 3
    end_row  = ws.max_row
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                test_data = str(ws.cell(row=row_no, column=4).value).strip()
                return test_data

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1



def test_status(test_id,test_step,test_status):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                ws.cell(row=row_no, column=9).value = test_status
                ws.cell(row=row_no, column=10).value = datetime.now()

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1
    wb.save("C:\\TM\\jatayu\\test_suite\\test_result.xlsx")



def AutomationException(test_id,test_step,exception):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                ws.cell(row=row_no, column=11).value = exception

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1
    wb.save("C:\\TM\\jatayu\\test_suite\\test_result.xlsx")


def WriteTestparameters(test_id,test_step,**kwargs):
    row_no = 3
    while row_no <= ws.max_row:
        if str(ws.cell(row=row_no, column=1).value).strip() == str(test_id):
            if str(ws.cell(row=row_no, column=6).value).strip() == str(test_step):
                col=12
                for key, value in kwargs.items():
                    ws.cell(row=row_no, column=col).value = key + "=" + value
                    col = col + 1

        elif ws.cell(row=row_no, column=1).value != None and int(str(ws.cell(row=row_no, column=1).value).strip()) > int(test_id):
            break
        row_no = row_no + 1
    wb.save("C:\\TM\\jatayu\\test_suite\\test_result.xlsx")

def get_element_repo(ElementName):
    row_no = 3
    end_row  = RepoSheet.max_row
    while row_no <= RepoSheet.max_row:
        if ElementName == str(RepoSheet.cell(row=row_no, column=2).value).strip():
            IdentifierType = str(RepoSheet.cell(row=row_no, column=3).value).strip()
            ElementType = str(RepoSheet.cell(row=row_no, column=4).value).strip()
            ElementIdentifier = str(RepoSheet.cell(row=row_no, column=5).value).strip()
            Action = str(RepoSheet.cell(row=row_no, column=6).value).strip()
            return IdentifierType, ElementType, ElementIdentifier, Action
        row_no = row_no + 1

wb.close()